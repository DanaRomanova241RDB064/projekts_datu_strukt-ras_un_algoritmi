from openpyxl import Workbook,load_workbook
import os 

wb=Workbook()
ws=wb.active

kd,ld,md,teo,eks={},{},{},{},{}

varianti={1:(kd,2,"kontroldarba"),2:(ld,3,"labaratorijas darba"),3:(md,4,"mājasdarba"),4:(teo,5,"teorijas testa"),5:(eks,6,"eksāmena")}

vert_veidi = [("Kontroldarbi", 2), ("Labaratorijas darbi", 3), ("Mājasdarbi", 4),("Teorijas testi", 5),("Eksāmens", 6)]

def int_kludas_apstrade(izvades_teksts):
    while True:
        try:
            return int(input(izvades_teksts))
        except ValueError:
            print("Lūdzu ievadiet veselo skaitli!")

def float_kludas_apstrade(izvades_teksts):
    while True:
        komats_uz_punktu=input(izvades_teksts).replace(",",".")
        try: 
            return float(komats_uz_punktu)
        except ValueError:
            print("Lūdzu ievadiet decimālskaitli ar punktu vai komatu!")

def faila_izveide(path="vertejumi.xlsx") :
    if not os.path.exists(path):
        wb=Workbook()
        ws=wb.active
        ws.title="Studiju kursu vērtējumi"
        
        kolonnu_nosaukumi=["Studiju kursa nosaukums","Kontroldarbu vidējais vērtējums", "Laboratorijas darbu vidējais vērtējums", "Mājasdarbu vidējais vērtējums","Teorijas testu vidējais vērtējums", "Eksāmena vērtējums", "Gala vidējais vērtējums"]
        for kolonna, nosaukums in enumerate(kolonnu_nosaukumi,start=1):
            ws.cell(row=1,column=kolonna,value=nosaukums)

        ws_vertejumi=wb.create_sheet("Vērtējumi")
        ws_vertejumi.append(["Studiju kurss","Vērtējuma veids","Vērtējumi"])
        print("Jauns fails tika veiksmīgi izveidots!")
        wb.save(path)
        wb.close()
    return load_workbook(path)

def nolasit_vertejumus(wb):
    vertejumu_lapa = wb["Vērtējumi"]
    
    for saraksti in (kd,ld,md,teo,eks):
        saraksti.clear()
    for kurss, vert_veids, vert in vertejumu_lapa.iter_rows(min_row=2,values_only=True):
        for saraksts, _, veids in varianti.values():
            if veids == vert_veids:
                saraksts.setdefault(kurss, []).append(vert)
                break

def pievienot_studiju_kursus(ws):
    skaits=int_kludas_apstrade("Cik studiju kursus tu vēlies pievienot? ")
    for _ in range(skaits):
        nosaukums=input("Ievadi studiju kursa nosaukumu: ")
        ws.append([nosaukums])

def vertejuma_veida_izvele():
    vertejuma_izvele=int_kludas_apstrade("Kāda vērtējuma veidu tu vēlies pievienot: \n1)Kontroldarbs\n2)Labaratorijas darbs\n3)Mājasdarbs\n4)Teorijas testa\n5)Eksāmens\n")
    if vertejuma_izvele in varianti:
        return varianti[vertejuma_izvele]
    else:
        print("Nederīga ievade!")
        return None

def izveleties_studiju_kursu(ws):
    for i,cell in enumerate(ws['A'][1:],start=1):
            print(i,") ",cell.value)
    izvele=int_kludas_apstrade("Izvēies studiju kursu, raksti tikai numuru: ")
    rinda=izvele+1
    kursa_nosaukums=ws.cell(row=izvele+1, column=1).value
    return rinda, kursa_nosaukums

def pievienot_vid_vert(ws):
    vertejuma_izvele=vertejuma_veida_izvele()
    vert_saraksts,kolonna,vertejuma_veids=vertejuma_izvele
    rinda, kursa_nosaukums= izveleties_studiju_kursu(ws)
    
    vert=float_kludas_apstrade("Ievadi "+vertejuma_veids+" vērtējumu: ")
    
    saraksts=vert_saraksts.setdefault(kursa_nosaukums, [])
    saraksts.append(vert)
    
    vid_vert=sum(saraksts)/len(saraksts)
    cell = ws.cell(row=rinda,column=kolonna,value=vid_vert)
    cell.number_format = '0.00'
    
    v_saraksts=ws.parent["Vērtējumi"]
    v_saraksts.append([kursa_nosaukums,vertejuma_veids,vert])
    
    print("Vidējais vērtējums kursam "+str(kursa_nosaukums)+" ir: "+str(vid_vert))

def studiju_kursu_ipatsvars_gala_vertejuma(vert_veidi):
    ipatsvari={}
    print("Ievadi izvēlētā studiju kursa īpatsvaru katra veida vērtējumam decimālskaitļos, piemēram, ja 30%, raksti 0,3, ja gala vērtējumu neveido kāds no vērtējumu veidiem, raksti 0")
    for veidi, _ in vert_veidi:
        ipatsvars = float_kludas_apstrade(veidi + ": ")
        ipatsvari[veidi] = float(ipatsvars)
    return ipatsvari

def aprekinat_gala_vertejumu(ws,rinda,ipatsvari):
    gala_vert=0
    for veidi, kolonna in vert_veidi:
        vertiba=ws.cell(row=rinda,column=kolonna).value or 0
        i_svars=ipatsvari.get(veidi, 0)
        gala_vert+=vertiba*i_svars
    return gala_vert

def saglabat_gala_vertejumu(ws, rinda, gala_vert):
    cell=ws.cell(row=rinda, column=7, value=gala_vert)
    cell.number_format='0.00'
    kursa_nosaukums= ws.cell(row=rinda, column=1).value
    print("Kursam "+kursa_nosaukums+" gala vērtējums ir: "+str(gala_vert))

def ierakstit_gala_vertejumu(ws):
    r, _=izveleties_studiju_kursu(ws)
    ipatsvari=studiju_kursu_ipatsvars_gala_vertejuma(vert_veidi)
    gala_vert=aprekinat_gala_vertejumu(ws,r,ipatsvari)
    saglabat_gala_vertejumu(ws, r, gala_vert)

def ievadit_laboto_vertejumu(veidi):
    return float_kludas_apstrade("Ievadi jauno "+veidi+" vērtējumu: ")

def ierakstit_jauno_vid_vert(ws, rinda, kolonna, vid_vert):
    cell=ws.cell(row=rinda, column=kolonna, value=vid_vert)
    cell.number_format='0.00'

def labot_pedejo_vertejumu(ws):
    r,kursa_nosaukums=izveleties_studiju_kursu(ws)
    vert_saraksts, kolonna, veidi=vertejuma_veida_izvele()
    if not vert_saraksts: return

    atsevisku_vertejumu_lapa=ws.parent["Vērtējumi"]
    for rinda in range (atsevisku_vertejumu_lapa.max_row, 1,-1):
        if(atsevisku_vertejumu_lapa.cell(row=rinda, column=1).value==kursa_nosaukums and atsevisku_vertejumu_lapa.cell(row=rinda,column=2).value==veidi):
            atsevisku_vertejumu_lapa.delete_rows(rinda,1)
            break

    saraksts = vert_saraksts.setdefault(kursa_nosaukums, [])
    if not saraksts:
        print("Šim kursam vēl nav vērtējumu, ko labot.")
        return
    
    jaunais_vert=ievadit_laboto_vertejumu(veidi)
    saraksts.pop()
    saraksts.append(jaunais_vert)
    vid_vert=sum(saraksts)/len(saraksts)
    
    ierakstit_jauno_vid_vert(ws,r,kolonna,vid_vert)
    ws.parent["Vērtējumi"].append([kursa_nosaukums, veidi, jaunais_vert])

while True:
    pirmas_darbibas_izvele = int_kludas_apstrade("Izvēlies ko tu vēlies darīt: \n 1)Izveidot jaunu failu, kurā glabāsies mani vērtējumi\n 2)Pievienot jaunu studiju kursu\n 3)Esošajam studiju kursam pievienot jaunu vērtējumu\n 4)Aprēķināt vidējo gala vērtējumu studiju kursiem\n 5)Labot pēdējo vērtējumu kādam kursam\n 6)Beigt darbu\n")
    
    if pirmas_darbibas_izvele==1:
       wb.close()
       wb=faila_izveide("vertejumi.xlsx")
       ws = wb["Studiju kursu vērtējumi"]
   
    elif pirmas_darbibas_izvele==2:
        wb=load_workbook("vertejumi.xlsx")
        nolasit_vertejumus(wb)
        ws = wb["Studiju kursu vērtējumi"]
        pievienot_studiju_kursus(ws)
        wb.save("vertejumi.xlsx")
    
    elif pirmas_darbibas_izvele==3:
        wb=load_workbook("vertejumi.xlsx")
        nolasit_vertejumus(wb)
        ws = wb["Studiju kursu vērtējumi"]
        pievienot_vid_vert(ws)
        wb.save("vertejumi.xlsx")
   
    elif pirmas_darbibas_izvele==4:
        wb=load_workbook("vertejumi.xlsx")
        nolasit_vertejumus(wb)
        ws = wb["Studiju kursu vērtējumi"]
        ierakstit_gala_vertejumu(ws)
        wb.save("vertejumi.xlsx")

    if pirmas_darbibas_izvele==5:
        wb=load_workbook("vertejumi.xlsx")
        nolasit_vertejumus(wb)
        ws = wb["Studiju kursu vērtējumi"]
        labot_pedejo_vertejumu(ws)
        wb.save("vertejumi.xlsx")
    
    if pirmas_darbibas_izvele==6:
        print("Darbs pabeigts!")
        break
    else:
        print("")

    wb.save("vertejumi.xlsx")
    wb.close
